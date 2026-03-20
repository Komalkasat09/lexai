"""
LexAI Human Evaluation Builder
===============================
Generates materials for human expert evaluation.

Outputs:
1. Anonymized evaluation forms (PDF/Excel)
2. Random 60-query subset (20 per system)
3. Shuffled responses (blinded)
4. 5-point Likert scales
5. Rating collection template

Usage:
    from evaluation.human_eval_builder import HumanEvalBuilder
    
    builder = HumanEvalBuilder()
    builder.generate_evaluation_materials(lexai_responses, baseline_responses)
"""

import random
import pandas as pd
from typing import Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Set seed for reproducibility
random.seed(42)


class HumanEvalBuilder:
    """
    Generates human evaluation materials.
    """
    
    def __init__(self, n_queries: int = 60, queries_per_system: int = 20):
        """
        Initialize human evaluation builder.
        
        Args:
            n_queries: Total number of queries for evaluation (default: 60)
            queries_per_system: Queries to evaluate per system (default: 20)
        """
        self.n_queries = n_queries
        self.queries_per_system = queries_per_system
        
        # 5 evaluation criteria (5-point Likert)
        self.criteria = [
            {
                "id": "q1",
                "name": "Legal Accuracy",
                "question": "How accurate is the legal information provided?",
                "scale": {
                    1: "Completely inaccurate",
                    2: "Mostly inaccurate",
                    3: "Somewhat accurate",
                    4: "Mostly accurate",
                    5: "Completely accurate"
                }
            },
            {
                "id": "q2",
                "name": "Citation Reliability",
                "question": "Are the legal citations (acts, sections, cases) correct and verifiable?",
                "scale": {
                    1: "No correct citations",
                    2: "Few correct citations",
                    3: "Some correct citations",
                    4: "Most citations correct",
                    5: "All citations correct"
                }
            },
            {
                "id": "q3",
                "name": "Practical Usefulness",
                "question": "How useful would this answer be for a practicing lawyer?",
                "scale": {
                    1: "Not useful at all",
                    2: "Slightly useful",
                    3: "Moderately useful",
                    4: "Very useful",
                    5: "Extremely useful"
                }
            },
            {
                "id": "q4",
                "name": "Trust Level",
                "question": "Would you trust this answer for legal advice?",
                "scale": {
                    1: "Completely distrust",
                    2: "Mostly distrust",
                    3: "Neutral",
                    4: "Mostly trust",
                    5: "Completely trust"
                }
            },
            {
                "id": "q5",
                "name": "Outdated Law Detection",
                "question": "Does the answer correctly handle outdated laws (IPC→BNS, overruled cases, amendments)?",
                "scale": {
                    1: "Completely misses",
                    2: "Mostly misses",
                    3: "Partially addresses",
                    4: "Mostly addresses",
                    5: "Fully addresses",
                    0: "Not applicable (query doesn't involve outdated law)"
                }
            }
        ]
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # STEP 1: Select Query Subset
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def select_query_subset(self, ground_truth: pd.DataFrame) -> List[int]:
        """
        Select random subset of queries for human evaluation.
        
        Ensures balanced category distribution.
        
        Args:
            ground_truth: Full ground truth dataset
            
        Returns:
            List of selected query indices
        """
        categories = ground_truth['category'].unique()
        queries_per_category = self.n_queries // len(categories)
        
        selected_indices = []
        
        for category in categories:
            category_queries = ground_truth[ground_truth['category'] == category]
            n_select = min(queries_per_category, len(category_queries))
            sampled = category_queries.sample(n=n_select, random_state=42)
            selected_indices.extend(sampled.index.tolist())
        
        # If we need more to reach n_queries, sample randomly
        if len(selected_indices) < self.n_queries:
            remaining = ground_truth.drop(selected_indices).sample(
                n=self.n_queries - len(selected_indices),
                random_state=42
            )
            selected_indices.extend(remaining.index.tolist())
        
        # Shuffle
        random.shuffle(selected_indices)
        
        return selected_indices[:self.n_queries]
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # STEP 2: Anonymize Responses
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def anonymize_responses(self, query_indices: List[int],
                           lexai_responses: List[Dict],
                           baseline_responses: Dict[str, List[Dict]]) -> pd.DataFrame:
        """
        Create anonymized evaluation dataset.
        
        Args:
            query_indices: Selected query indices
            lexai_responses: LexAI responses
            baseline_responses: Dict of {baseline_name: responses}
            
        Returns:
            DataFrame with anonymized responses
        """
        evaluation_data = []
        
        systems = ['LexAI'] + list(baseline_responses.keys())
        
        # For each query, select one system's response
        for idx in query_indices:
            # Randomly select which system to evaluate for this query
            selected_system = random.choice(systems)
            
            if selected_system == 'LexAI':
                response = lexai_responses[idx]
            else:
                response = baseline_responses[selected_system][idx]
            
            evaluation_data.append({
                'eval_id': len(evaluation_data) + 1,
                'query_index': idx,
                'system_id': f"System_{random.randint(1000, 9999)}",  # Anonymous ID
                'actual_system': selected_system,  # Hidden from evaluators
                'response': response.get('answer', ''),
                'citations': self._extract_citations_display(response),
                'confidence': response.get('confidence', 'N/A')
            })
        
        return pd.DataFrame(evaluation_data)
    
    def _extract_citations_display(self, response: Dict) -> str:
        """Extract citations for display."""
        citations = []
        
        if 'structured_response' in response:
            struct = response['structured_response']
            if 'act_cited' in struct:
                citations.append(f"Act: {struct['act_cited']}")
            if 'section_cited' in struct:
                citations.append(f"Section: {struct['section_cited']}")
            if 'case_citations' in struct:
                citations.append(f"Cases: {', '.join(struct['case_citations'][:3])}")
        
        return "; ".join(citations) if citations else "No explicit citations"
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # STEP 3: Generate Excel Evaluation Form
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def generate_excel_form(self, eval_data: pd.DataFrame, ground_truth: pd.DataFrame,
                           output_path: str = "evaluation/human_eval_form.xlsx"):
        """
        Generate Excel evaluation form for human evaluators.
        
        Args:
            eval_data: Anonymized evaluation data
            ground_truth: Ground truth dataset
            output_path: Output file path
        """
        wb = openpyxl.Workbook()
        
        # Sheet 1: Instructions
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"
        self._create_instructions_sheet(ws_instructions)
        
        # Sheet 2: Evaluation Form
        ws_eval = wb.create_sheet("Evaluation Form")
        self._create_evaluation_sheet(ws_eval, eval_data, ground_truth)
        
        # Sheet 3: Rating Scale Reference
        ws_scale = wb.create_sheet("Rating Scale")
        self._create_scale_sheet(ws_scale)
        
        # Save
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"  ✓ Generated: {output_path}")
    
    def _create_instructions_sheet(self, ws):
        """Create instructions sheet."""
        ws['A1'] = "LexAI Human Evaluation - Instructions"
        ws['A1'].font = Font(size=16, bold=True)
        
        instructions = [
            "",
            "Thank you for participating in this evaluation!",
            "",
            "TASK:",
            "You will evaluate legal AI system responses across 60 queries.",
            "For each query-response pair, rate the response on 5 criteria using a 5-point scale.",
            "",
            "CRITERIA:",
            "1. Legal Accuracy - Correctness of legal information",
            "2. Citation Reliability - Accuracy of legal citations",
            "3. Practical Usefulness - Value for practicing lawyers",
            "4. Trust Level - Whether you would trust this answer",
            "5. Outdated Law Detection - Handling of IPC→BNS transitions, amendments, overruled cases",
            "",
            "RATING SCALE:",
            "1 = Lowest (e.g., Completely Inaccurate)",
            "2 = Low",
            "3 = Moderate",
            "4 = High",
            "5 = Highest (e.g., Completely Accurate)",
            "0 = Not Applicable (only for criterion 5)",
            "",
            "INSTRUCTIONS:",
            "• The responses are anonymized - you do not know which AI system generated each response",
            "• Please evaluate each response independently based on your legal expertise",
            "• Refer to the 'Rating Scale' sheet for detailed scale descriptions",
            "• Enter your ratings in the colored columns in 'Evaluation Form' sheet",
            "• Optional: Add comments in the 'Comments' column for interesting cases",
            "",
            "ESTIMATED TIME: 2-3 hours",
            "",
            "Your evaluator ID: _______",
            "Date: _______"
        ]
        
        for i, line in enumerate(instructions, start=2):
            ws[f'A{i}'] = line
    
    def _create_evaluation_sheet(self, ws, eval_data: pd.DataFrame, ground_truth: pd.DataFrame):
        """Create main evaluation sheet."""
        # Headers
        headers = [
            "Eval ID", "Query", "System Response", "Citations",
            "Q1: Legal Accuracy\n(1-5)", 
            "Q2: Citation Reliability\n(1-5)",
            "Q3: Practical Usefulness\n(1-5)",
            "Q4: Trust Level\n(1-5)",
            "Q5: Outdated Law\n(0-5)",
            "Comments (Optional)"
        ]
        
        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        
        # Data rows
        for row_idx, (_, row) in enumerate(eval_data.iterrows(), start=2):
            gt = ground_truth.iloc[row['query_index']]
            
            ws.cell(row=row_idx, column=1, value=row['eval_id'])
            ws.cell(row=row_idx, column=2, value=gt['query'])
            ws.cell(row=row_idx, column=3, value=row['response'])
            ws.cell(row=row_idx, column=4, value=row['citations'])
            
            # Rating columns (colored yellow for input)
            for col_idx in range(5, 10):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Comments column
            ws.cell(row=row_idx, column=10, value="")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 40
        for col in ['E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['J'].width = 30
    
    def _create_scale_sheet(self, ws):
        """Create rating scale reference sheet."""
        ws['A1'] = "Rating Scale Reference"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        for criterion in self.criteria:
            ws[f'A{row}'] = criterion['name']
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            ws[f'A{row}'] = criterion['question']
            ws[f'A{row}'].font = Font(italic=True)
            row += 1
            
            for score, description in sorted(criterion['scale'].items()):
                ws[f'A{row}'] = f"  {score} = {description}"
                row += 1
            
            row += 1  # Blank line
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # STEP 4: Generate Results Collection Template
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def load_evaluator_ratings(self, eval_form_path: str) -> pd.DataFrame:
        """
        Load completed evaluation form.
        
        Args:
            eval_form_path: Path to completed Excel form
            
        Returns:
            DataFrame with ratings
        """
        df = pd.read_excel(eval_form_path, sheet_name="Evaluation Form")
        
        # Validate completeness
        rating_cols = [col for col in df.columns if col.startswith('Q')]
        missing = df[rating_cols].isna().sum().sum()
        
        if missing > 0:
            print(f"  ⚠ Warning: {missing} missing ratings found")
        
        return df
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # MASTER FUNCTION
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def generate_evaluation_materials(self, ground_truth: pd.DataFrame,
                                     lexai_responses: List[Dict],
                                     baseline_responses: Dict[str, List[Dict]],
                                     output_dir: str = "evaluation") -> Dict:
        """
        Generate complete human evaluation materials.
        
        Args:
            ground_truth: Ground truth dataset
            lexai_responses: LexAI responses
            baseline_responses: Baseline responses dict
            output_dir: Output directory
            
        Returns:
            Dictionary with paths and metadata
        """
        print("Generating human evaluation materials...")
        
        # Step 1: Select queries
        print("  1/4 Selecting query subset...")
        query_indices = self.select_query_subset(ground_truth)
        
        # Step 2: Anonymize responses
        print("  2/4 Anonymizing responses...")
        eval_data = self.anonymize_responses(query_indices, lexai_responses, baseline_responses)
        
        # Step 3: Generate Excel form
        print("  3/4 Generating Excel evaluation form...")
        excel_path = os.path.join(output_dir, "human_eval_form.xlsx")
        self.generate_excel_form(eval_data, ground_truth, excel_path)
        
        # Step 4: Save metadata (for later de-anonymization)
        print("  4/4 Saving metadata...")
        metadata = eval_data[['eval_id', 'query_index', 'actual_system']].copy()
        metadata_path = os.path.join(output_dir, "human_eval_metadata.xlsx")
        metadata.to_excel(metadata_path, index=False)
        print(f"  ✓ Generated: {metadata_path}")
        
        print("  ✓ Human evaluation materials complete")
        
        return {
            "excel_form": excel_path,
            "metadata": metadata_path,
            "n_queries": len(query_indices),
            "query_indices": query_indices,
            "instructions": "Share 'human_eval_form.xlsx' with evaluators. Keep 'metadata' private."
        }


def demo():
    """Demo function."""
    print("Human Evaluation Builder Demo")
    print("=" * 60)
    print("\nThis module generates human evaluation materials:")
    print("  • Selects 60-query random subset")
    print("  • Anonymizes system responses")
    print("  • Generates Excel evaluation form")
    print("  • 5 evaluation criteria × 5-point Likert scales")
    print("  • Supports multiple evaluators for inter-rater reliability")
    print("\nOutput: human_eval_form.xlsx (for evaluators)")
    print("        human_eval_metadata.xlsx (for de-anonymization)")


if __name__ == "__main__":
    demo()
